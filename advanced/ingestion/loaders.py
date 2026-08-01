"""
Document loaders for PDF, DOCX, TXT, XLSX.
Each loader returns a list of LangChain Document objects with source metadata.
Errors are caught and surfaced clearly — never silently swallowed into empty results,
per the RAG chatbot's explicit error-handling boundary (Rules.md).
"""

from pathlib import Path
from langchain_core.documents import Document
from langchain_community.document_loaders import (
    PyPDFLoader, Docx2txtLoader, TextLoader
)
import openpyxl


class UnsupportedFileTypeError(Exception):
    """Raised when a file extension isn't in the supported set."""
    pass


class DocumentLoadError(Exception):
    """Raised when a supported file fails to load (corrupt, empty, unreadable)."""
    pass


def load_pdf(filepath: Path) -> list[Document]:
    try:
        loader = PyPDFLoader(str(filepath))
        docs = loader.load()
        if not docs or all(not d.page_content.strip() for d in docs):
            raise DocumentLoadError(
                f"'{filepath.name}' loaded but contains no extractable text "
                "(may be a scanned/image-only PDF)."
            )
        return docs
    except DocumentLoadError:
        raise
    except Exception as e:
        raise DocumentLoadError(f"Failed to load PDF '{filepath.name}': {e}") from e


def load_docx(filepath: Path) -> list[Document]:
    try:
        loader = Docx2txtLoader(str(filepath))
        docs = loader.load()
        if not docs or all(not d.page_content.strip() for d in docs):
            raise DocumentLoadError(f"'{filepath.name}' loaded but contains no text.")
        return docs
    except DocumentLoadError:
        raise
    except Exception as e:
        raise DocumentLoadError(f"Failed to load DOCX '{filepath.name}': {e}") from e


def load_txt(filepath: Path) -> list[Document]:
    try:
        loader = TextLoader(str(filepath), encoding="utf-8")
        docs = loader.load()
        if not docs or all(not d.page_content.strip() for d in docs):
            raise DocumentLoadError(f"'{filepath.name}' is empty.")
        return docs
    except DocumentLoadError:
        raise
    except UnicodeDecodeError as e:
        raise DocumentLoadError(
            f"'{filepath.name}' isn't valid UTF-8 text — check the file encoding."
        ) from e
    except Exception as e:
        raise DocumentLoadError(f"Failed to load TXT '{filepath.name}': {e}") from e


def load_xlsx(filepath: Path) -> list[Document]:
    """
    Loads an XLSX file, converting each sheet to a text representation.
    No off-the-shelf LangChain XLSX loader is reliable enough across sheet
    layouts, so this is hand-rolled using openpyxl directly.
    """
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        docs = []

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            rows_text = []

            for row in sheet.iter_rows(values_only=True):
                if any(cell is not None for cell in row):
                    row_text = " | ".join(
                        str(cell) if cell is not None else "" for cell in row
                    )
                    rows_text.append(row_text)

            if rows_text:
                sheet_content = f"Sheet: {sheet_name}\n" + "\n".join(rows_text)
                docs.append(Document(
                    page_content=sheet_content,
                    metadata={"source": str(filepath), "sheet": sheet_name}
                ))

        if not docs:
            raise DocumentLoadError(f"'{filepath.name}' has no non-empty sheets.")

        return docs
    except DocumentLoadError:
        raise
    except Exception as e:
        raise DocumentLoadError(f"Failed to load XLSX '{filepath.name}': {e}") from e


LOADER_MAP = {
    ".pdf": load_pdf,
    ".docx": load_docx,
    ".txt": load_txt,
    ".xlsx": load_xlsx,
}


def load_document(filepath: str | Path) -> list[Document]:
    """
    Main entry point — dispatches to the correct loader based on extension.

    Args:
        filepath: path to the file to load

    Returns:
        list[Document]: LangChain documents with source metadata attached

    Raises:
        UnsupportedFileTypeError: if the extension isn't supported
        DocumentLoadError: if loading fails for any reason
        FileNotFoundError: if the file doesn't exist
    """
    filepath = Path(filepath)

    if not filepath.exists():
        raise FileNotFoundError(f"File not found: {filepath}")

    ext = filepath.suffix.lower()

    if ext not in LOADER_MAP:
        raise UnsupportedFileTypeError(
            f"'{ext}' is not a supported file type. "
            f"Supported types: {', '.join(sorted(LOADER_MAP.keys()))}"
        )

    docs = LOADER_MAP[ext](filepath)

    # Ensure every doc carries a consistent, human-readable source filename
    for doc in docs:
        doc.metadata["source_file"] = filepath.name

    return docs